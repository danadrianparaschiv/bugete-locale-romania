"""Deterministic ingestion of official municipal Excel budget sources.

The public corpus prefers a native workbook when the municipality publishes
one.  This module converts both OOXML and Excel 97-2003 sheets into the same
page payload contract used by PDF extraction, so assembly, validation,
publication, analytics and auditing remain shared.
"""

from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any

from .assemble import assemble
from .extract.scanned import map_payload
from .layouts.common import fold
from .model import ConversionResult
from .parsing import normalize_indicator_code
from .validate import validate

SUPPORTED_SUFFIXES = frozenset({".xls", ".xlsx"})
_STANDARD_HEADER = re.compile(r"denumirea indicatorului bugetar")
_COMPARATIVE_TITLE = re.compile(r"proiectul bugetului local", re.I)
_CONSOLIDATED_TITLE = re.compile(r"bugetul general consolidat", re.I)
_CAPITOL = re.compile(r"\bcap\.?\s*(\d{2})\b", re.I)
_PRINTED_CODE = re.compile(r"(?<!\d)(\d{4,10})(?!\d)")


class _PayloadStore:
    """Small RunStore-compatible view used by ``assemble``."""

    def __init__(self, payloads: dict[int, dict[str, Any]]):
        self.payloads = payloads

    def get(self, stage: str, page: int):
        return self.payloads.get(page) if stage == "extract" else None


def _cell_text(value: Any, number_format: str | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        fmt = (number_format or "").strip()
        # Indicator codes are occasionally stored as 0.01 with Excel format
        # ``00.00``.  Preserve the displayed leading zero instead of turning
        # the budget total code into the unrelated string ``0.01``.
        if re.fullmatch(r"0{1,4}(?:\.0{1,4})+", fmt):
            decimals = len(fmt.rsplit(".", 1)[1])
            width = len(fmt)
            return f"{float(value):0{width}.{decimals}f}"
        if float(value).is_integer():
            return str(int(value))
        return format(Decimal(str(value)), "f")
    return str(value).strip()


def _xlsx_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for name in workbook.sheetnames:
            sheet = workbook[name]
            rows = [
                [_cell_text(cell.value, cell.number_format) for cell in row]
                for row in sheet.iter_rows()
            ]
            sheets.append((name, rows))
        return sheets
    finally:
        workbook.close()


def _xls_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        sheets = []
        for index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(index)
            rows = [
                [_cell_text(sheet.cell_value(row, column)) for column in range(sheet.ncols)]
                for row in range(sheet.nrows)
            ]
            sheets.append((sheet.name, rows))
        return sheets
    finally:
        workbook.release_resources()


def read_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    """Read displayed values without mutating or recalculating the source."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _xlsx_sheets(path)
    if suffix == ".xls":
        return _xls_sheets(path)
    raise ValueError(f"unsupported native workbook format: {path.suffix}")


def _metadata_text(sheet_name: str, grid: list[list[str]]) -> str:
    cells = [cell for row in grid[:12] for cell in row if cell.strip()]
    return " ".join([sheet_name, *cells])


def _trim_grid(grid: list[list[str]]) -> list[list[str]]:
    """Drop formatting-only trailing rows/columns from exported worksheets."""
    rows = [list(row) for row in grid]
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    width = max(
        (index + 1 for row in rows for index, cell in enumerate(row) if cell.strip()),
        default=0,
    )
    return [row[:width] for row in rows]


def _source_is_lei(text: str) -> bool:
    normalized = fold(text)
    return bool(re.search(r"(?:^|[ (\-])lei(?:$|[ )\-])", normalized)) and not bool(
        re.search(r"mii\s+lei", normalized)
    )


def _normalize_to_mii_lei(lines: list[dict[str, Any]], source_is_lei: bool) -> None:
    divisor = Decimal(1000) if source_is_lei else Decimal(1)
    for line in lines:
        values = line.get("values") or {}
        if divisor != 1:
            line["values"] = {
                column: str(Decimal(str(value)) / divisor)
                if value != "X" else value
                for column, value in values.items()
            }
        line["source"] = "native_excel"
        if line.get("values"):
            line["value_sources"] = {
                column: "native_excel" for column in line["values"]
            }


def _standard_payload(
    sheet_name: str, grid: list[list[str]], budget_year: int
) -> dict[str, Any] | None:
    grid = _trim_grid(grid)
    header = next(
        (
            index
            for index, row in enumerate(grid)
            if _STANDARD_HEADER.search(fold(" ".join(row)))
        ),
        None,
    )
    if header is None:
        return None
    text = _metadata_text(sheet_name, grid)
    canonical_title = "BUGETUL LOCAL DETALIAT " if "02a" in fold(text) else ""
    payload = map_payload(
        {"tables_raw": [grid[header:]], "text": canonical_title + text.upper()},
        budget_year=budget_year,
    )
    payload["layout"] = "native_excel_standard"
    payload["source_format"] = "native_excel"
    payload["source_unit"] = "lei" if _source_is_lei(text) else "mii lei"
    payload["output_unit"] = "mii lei"
    _normalize_to_mii_lei(payload["lines"], payload["source_unit"] == "lei")
    return payload


def _registry_code(raw_digits: str, registry) -> str | None:
    direct = normalize_indicator_code(raw_digits)
    candidates = [direct]
    # Some locally-authored sheets omit the funding-source pair ``02``:
    # 031800 means 03.02.18 and 0405 means 04.02.05.
    candidates.append(normalize_indicator_code(raw_digits[:2] + "02" + raw_digits[2:]))
    if raw_digits.endswith("00"):
        candidates.extend([
            normalize_indicator_code(raw_digits[:-2]),
            normalize_indicator_code(raw_digits[:2] + "02" + raw_digits[2:-2]),
        ])
    for candidate in candidates:
        if candidate and registry.exists(candidate):
            return candidate
    return direct


def comparative_payload(
    sheet_name: str,
    grid: list[list[str]],
    budget_year: int,
    registry,
) -> dict[str, Any] | None:
    """Map the locally-authored Satu Mare comparison sheet.

    The official sheet compares several historic budgets across 24 columns.
    Only the explicitly labelled current-year budget column is ingested; code
    completion is accepted only against the historical official registry.
    """
    text = _metadata_text(sheet_name, grid)
    if not _COMPARATIVE_TITLE.search(text):
        return None
    header = next(
        (
            index
            for index, row in enumerate(grid)
            if any(f"buget {budget_year}" in fold(cell) for cell in row)
        ),
        None,
    )
    if header is None:
        return None
    year_column = next(
        index
        for index, cell in enumerate(grid[header])
        if f"buget {budget_year}" in fold(cell)
    )
    total_revenue = next(
        (
            row[year_column]
            for row in grid[header + 1 :]
            if row and fold(row[0]).startswith("total venituri")
            and year_column < len(row) and row[year_column]
        ),
        "",
    )
    total_expense = next(
        (
            row[year_column]
            for row in grid[header + 1 :]
            if row and fold(row[0]).startswith("total cheltuieli")
            and year_column < len(row) and row[year_column]
        ),
        "",
    )
    lines: list[dict[str, Any]] = []

    def add_marker(name: str, code: str, raw_value: str) -> None:
        values = {}
        if raw_value:
            values[f"total_{budget_year}"] = str(Decimal(raw_value) / Decimal(1000))
        lines.append({
            "raw_code": code.replace(".", ""),
            "code": code,
            "func_code": None,
            "name": name,
            "row_no": None,
            "section": "FUNCTIONARE",
            "year": budget_year,
            "values": values,
            "source": "native_excel",
            "value_sources": {column: "native_excel" for column in values},
        })

    add_marker("TOTAL VENITURI", "00.01", total_revenue)
    in_expenses = False
    expense_marker_added = False
    for row in grid[header + 1 :]:
        if not row or not row[0].strip():
            continue
        name = " ".join(row[0].split())
        normalized_name = fold(name)
        if normalized_name == "cheltuieli":
            in_expenses = True
            if not expense_marker_added:
                add_marker("TOTAL CHELTUIELI", "50.02", total_expense)
                expense_marker_added = True
            continue
        if normalized_name.startswith(("total venituri", "total cheltuieli")):
            continue
        if year_column >= len(row) or not row[year_column].strip():
            continue
        raw_value = row[year_column].strip()
        try:
            value = Decimal(raw_value) / Decimal(1000)
        except Exception:
            continue

        code = None
        match = _CAPITOL.search(name) if in_expenses else None
        if match:
            code = f"{match.group(1)}.02"
        elif not in_expenses:
            candidates = _PRINTED_CODE.findall(name)
            if candidates:
                code = _registry_code(candidates[-1], registry)
        if not code:
            # The source does not identify an economic code for these local
            # explanatory rows.  Omitting them is safer than inventing a
            # classification; the public quality contract does not claim
            # recall for facts absent from the normalized result.
            continue
        clean_name = re.sub(r"[-–—]?\s*\d{4,10}\s*[-–—]?\s*$", "", name).strip()
        lines.append({
            "raw_code": code.replace(".", ""),
            "code": code,
            "func_code": None,
            "name": clean_name or name,
            "row_no": None,
            "section": "FUNCTIONARE",
            "year": budget_year,
            "values": {f"total_{budget_year}": str(value)},
            "source": "native_excel",
            "value_sources": {f"total_{budget_year}": "native_excel"},
        })
    return {
        "lines": lines,
        "text": text.upper(),
        "layout": "native_excel_comparative",
        "n_tables": 1,
        "n_numeric_cells": len(lines),
        "budget_year": budget_year,
        "source_format": "native_excel",
        "source_unit": "lei",
        "output_unit": "mii lei",
        "mapping_stats": {
            "source_value_cells": len(lines),
            "mapped_value_cells": len(lines),
            "coded_value_lines": len(lines),
            "value_lines": len(lines),
            "cell_issues": 0,
        },
    }


_CONSOLIDATED_REVENUES = {
    "impozit pe profit": "01.02",
    "impozitul pe veniturile din transferul": "03.02.18",
    "cote si sume defalcate din impozitul pe venit": "04.02",
    "alte impozite pe venit": "05.02",
    "impozite si taxe pe proprietate": "07.02",
    "sume defalcate din tva": "11.02",
    "subventii de la bugetul de stat": "42.02",
    "subventii de la alte administratii": "43.02",
    "sume primite de la ue": "48.02",
}

_CONSOLIDATED_CHAPTERS = {
    "autoritati publice si actiuni externe": "51.02",
    "alte servicii publice generale": "54.02",
    "tranzactii privind datoria publica si imprumuturi": "55.02",
    "transferuri cu caracter general": "56.02",
    "aparare": "60.02",
    "ordine publica si siguranta nationala": "61.02",
    "invatamant": "65.02",
    "sanatate": "66.02",
    "cultura, recreere si religie": "67.02",
    "asigurari si asistenta sociala": "68.02",
    "locuinte, servicii si dezvoltare publica": "70.02",
    "protectia mediului": "74.02",
    "actiuni generale economice": "80.02",
    "combustibili si energie": "81.02",
    "agricultura, silvicultura, piscicultura si vanatoare": "83.02",
    "transporturi": "84.02",
    "alte actiuni economice": "87.02",
}


def consolidated_payload(
    sheet_name: str, grid: list[list[str]], budget_year: int
) -> dict[str, Any] | None:
    """Map a general-consolidated summary's local-budget column.

    Printed ``Cod rând`` values are form ordinals, not nomenclator codes. The
    chapter identities therefore come from a small, explicit name mapping;
    only names represented in the official functional classification enter
    the normalized result.
    """
    text = _metadata_text(sheet_name, grid)
    if not _CONSOLIDATED_TITLE.search(fold(text)):
        return None
    header = next(
        (
            row
            for row in grid
            if any(
                " ".join(fold(cell).split()).startswith("bugetul local")
                for cell in row
            )
        ),
        None,
    )
    if header is None:
        return None
    local_column = next(
        index
        for index, cell in enumerate(header)
        if " ".join(fold(cell).split()).startswith("bugetul local")
    )
    lines: list[dict[str, Any]] = []

    def value_at(row: list[str]) -> str | None:
        if local_column >= len(row) or not row[local_column].strip():
            return None
        try:
            return str(Decimal(row[local_column].strip()) / Decimal(1000))
        except Exception:
            return None

    def add_line(
        name: str,
        code: str | None,
        value: str,
        section: str,
        source: str = "native_excel",
    ) -> None:
        values = {f"total_{budget_year}": value}
        lines.append({
            "raw_code": code.replace(".", "") if code else None,
            "code": code,
            "func_code": None,
            "name": name,
            "row_no": None,
            "section": section,
            "year": budget_year,
            "values": values,
            "source": source,
            "value_sources": {column: source for column in values},
        })

    chapter_region = False
    for index, row in enumerate(grid):
        if not row or not row[0].strip():
            continue
        name = " ".join(row[0].split())
        normalized = fold(name)
        value = value_at(row)
        if normalized.startswith("venituri total") and value is not None:
            add_line("TOTAL VENITURI", "00.01", value, "TOTAL")
        elif normalized.startswith("cheltuieli - total") and value is not None:
            add_line("TOTAL CHELTUIELI", "49.02", value, "TOTAL")
        elif normalized.startswith("pe capitole"):
            chapter_region = True
        elif not chapter_region and value is not None:
            matched = next(
                (
                    code
                    for prefix, code in _CONSOLIDATED_REVENUES.items()
                    if normalized.startswith(prefix)
                ),
                None,
            )
            if matched:
                add_line(name, matched, value, "TOTAL")
        elif chapter_region:
            matched = next(
                (
                    code
                    for prefix, code in _CONSOLIDATED_CHAPTERS.items()
                    if normalized.startswith(prefix)
                ),
                None,
            )
            if not matched or value is None:
                continue
            add_line(name, matched, value, "TOTAL")
            for offset, section in ((1, "FUNCTIONARE"), (2, "DEZVOLTARE")):
                if index + offset >= len(grid):
                    continue
                section_row = grid[index + offset]
                if not section_row or not fold(section_row[0]).startswith("sectiunea"):
                    continue
                section_value = value_at(section_row)
                if section_value is not None:
                    add_line(name, matched, section_value, section)

    if not any(line["code"] == "00.01" for line in lines):
        return None
    printed_expense_total = next(
        (
            Decimal(line["values"][f"total_{budget_year}"])
            for line in lines
            if line["code"] == "49.02" and line["section"] == "TOTAL"
        ),
        None,
    )
    section_sums = {
        section: sum(
            (
                Decimal(line["values"][f"total_{budget_year}"])
                for line in lines
                if line["code"] in _CONSOLIDATED_CHAPTERS.values()
                and line["section"] == section
            ),
            Decimal(0),
        )
        for section in ("FUNCTIONARE", "DEZVOLTARE")
    }
    # The source prints section values per chapter but no grand section row.
    # Publish their sums only when the two independently close to the printed
    # expense total; provenance makes the derivation explicit per cell.
    if (
        printed_expense_total is not None
        and all(section_sums.values())
        and sum(section_sums.values()) == printed_expense_total
    ):
        for section, value in section_sums.items():
            add_line(
                "TOTAL CHELTUIELI",
                "49.02",
                str(value),
                section,
                source="native_excel:derived_sum",
            )
    return {
        "lines": lines,
        "text": f"BUGETUL LOCAL DETALIAT {text}".upper(),
        "layout": "native_excel_consolidated",
        "n_tables": 1,
        "n_numeric_cells": len(lines),
        "budget_year": budget_year,
        "source_format": "native_excel",
        "source_unit": "lei",
        "output_unit": "mii lei",
        "mapping_stats": {
            "source_value_cells": len(lines),
            "mapped_value_cells": len(lines),
            "coded_value_lines": sum(line["code"] is not None for line in lines),
            "value_lines": len(lines),
            "cell_issues": 0,
        },
    }


def workbook_payloads(
    path: Path, budget_year: int, registry
) -> dict[int, dict[str, Any]]:
    """Map each worksheet to the shared page-payload contract.

    Keeping this stage public lets the offline annotation inventory display and
    score native Excel sheets with the same source-unit boundary used by the
    converter.  Reading is deterministic and never mutates the official file.
    """
    payloads: dict[int, dict[str, Any]] = {}
    for page, (sheet_name, grid) in enumerate(read_sheets(path), 1):
        payload = _standard_payload(sheet_name, grid, budget_year)
        if payload is None:
            payload = consolidated_payload(sheet_name, grid, budget_year)
        if payload is None:
            payload = comparative_payload(sheet_name, grid, budget_year, registry)
        if payload is None:
            payload = {
                "lines": [],
                "text": _metadata_text(sheet_name, grid).upper(),
                "layout": "native_excel_metadata",
                "source_format": "native_excel",
                "source_unit": None,
                "output_unit": "mii lei",
            }
        payloads[page] = payload
    return payloads


def convert_workbook(path: Path, budget_year: int, registry) -> ConversionResult:
    """Convert an official native workbook through shared assembly/validation."""
    payloads = workbook_payloads(path, budget_year, registry)

    pages = list(payloads)
    documents = assemble(_PayloadStore(payloads), pages, registry)
    result = ConversionResult(
        pdf=path.name,
        documents=documents,
        pages_expected=len(pages),
        pages_selected=pages,
        pages_processed=pages,
    )
    validate(result, registry)
    return result
