"""Forexebug quarterly execution reports (FXB-EXB-901) -> validated snapshots.

Unlike the budget PDFs, these are already-structured Excel exports from the
Ministry of Finance, so there is no extraction risk: the work is mapping
Forexebug's 6-digit codes onto the classification this project already
validates against (Ordinul MFP 1954/2005) and cross-checking the report's
own totals.

Two conventions differ from the rest of the corpus and are normalized here:

* Forexebug prints values in **lei**; everything else in this project is in
  **mii lei**, so parsed values are divided by 1000.
* Forexebug omits the budget suffix from functional/revenue codes (``510103``)
  and carries it instead in the funding-source column: sources A–D belong to
  the local budget (``.02``), E–G to the self-financed institutions (``.10``).
  So ``510103`` + source ``A`` becomes ``51.02.01.03``.
"""

from __future__ import annotations

import json
import logging
import re
from datetime import date, timedelta
from pathlib import Path

from pydantic import BaseModel, Field

from .nomenclator import Registry

log = logging.getLogger("bgc.execution")

SCHEMA_VERSION = 1
LEI_PER_MIE = 1000

# funding-source letter -> which budget document the line belongs to
SOURCE_BUDGET = {
    "A": "local",        # integral de la buget
    "B": "local",        # credite externe
    "C": "local",        # credite interne
    "D": "local",        # fonduri externe nerambursabile
    "E": "own_revenue",  # activități finanțate integral din venituri proprii
    "F": "own_revenue",  # integral venituri proprii
    "G": "own_revenue",  # venituri proprii și subvenții
}
BUDGET_SUFFIX = {"local": "02", "own_revenue": "10"}
SECTION = {"F": "FUNCTIONARE", "D": "DEZVOLTARE"}

_DATE_RE = re.compile(r"LA DATA:\s*(\d{2})-([A-Z]{3})-(\d{2})", re.I)
_CIF_RE = re.compile(r"Cod Fiscal IP:\s*(\d+)\s*Denumire IP\s*:\s*(.+?)\s*$", re.I)
_MONTHS = {m: i for i, m in enumerate(
    ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], 1)}


class ExecutionLine(BaseModel):
    kind: str  # revenue | expense_functional
    code: str  # dotted functional/revenue code, budget suffix restored
    econ_code: str | None = None  # dotted economic code (expense lines)
    name: str
    section: str | None = None  # FUNCTIONARE | DEZVOLTARE
    budget: str  # local | own_revenue
    source_class: str  # A..G
    value: float  # mii lei, cumulative since 1 January
    known_code: bool = True  # False when the registry does not enumerate it


class ExecutionReport(BaseModel):
    """One parsed Forexebug workbook."""

    report_type: str = "FXB-EXB-901"
    report_date: str | None = None  # ISO
    entity_cif: str | None = None
    entity_name: str | None = None
    quarter: int | None = None
    lines: list[ExecutionLine] = Field(default_factory=list)
    printed_totals: dict[str, float] = Field(default_factory=dict)  # mii lei
    issues: list[str] = Field(default_factory=list)

    def total(self, kind: str, budget: str = "local") -> float:
        return sum(ln.value for ln in self.lines
                   if ln.kind == kind and ln.budget == budget)


def _iso_date(text: str) -> str | None:
    m = _DATE_RE.search(text)
    if not m:
        return None
    day, mon, yy = m.groups()
    month = _MONTHS.get(mon.upper())
    if not month:
        return None
    return f"20{yy}-{month:02d}-{int(day):02d}"


def dotted_code(raw: str | int | float, kind: str, budget: str) -> str | None:
    """``510103`` + local -> ``51.02.01.03``; economic codes keep no suffix."""
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None
    s = digits.zfill(6)[:6]
    head, rest = s[0:2], [s[2:4], s[4:6]]
    tail = [p for p in rest if p != "00"]
    if kind == "expense_economic":
        return ".".join([head, *tail])
    return ".".join([head, BUDGET_SUFFIX[budget], *tail])


def parse_report(path: Path, quarter: int | None = None,
                 registry: Registry | None = None) -> ExecutionReport:
    """Parse one Forexebug workbook; values normalized to mii lei."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    rep = ExecutionReport(quarter=quarter)
    header_at = None
    for i, row in enumerate(rows[:40]):
        text = " ".join(str(c) for c in row if c is not None)
        if rep.report_date is None:
            rep.report_date = _iso_date(text)
        m = _CIF_RE.search(text)
        if m and rep.entity_cif is None:
            rep.entity_cif, rep.entity_name = m.group(1), m.group(2).strip()
        if row and row[0] and str(row[0]).strip() == "Tip Indicator":
            header_at = i
            break
    if header_at is None:
        rep.issues.append("antetul tabelului nu a fost găsit")
        return rep

    # column positions differ between the report's two sheets (one carries a
    # spacer column), so bind them by header label instead of fixed indices
    col: dict[str, int] = {}
    for j, cell in enumerate(rows[header_at]):
        label = str(cell or "").strip().lower()
        if label.startswith("clasificatie functionala"):
            col["func_desc" if "descriere" in label else "func"] = j
        elif label.startswith("clasificatie economica"):
            col["econ_desc" if "descriere" in label else "econ"] = j
        elif label.startswith("sectiune"):
            col["section"] = j
        elif label.startswith("executie"):
            col["value"] = j
        elif label.startswith("sursa"):
            col["source"] = j
    missing = {"func", "value", "source"} - set(col)
    if missing:
        rep.issues.append(f"coloane lipsă în antet: {sorted(missing)}")
        return rep

    def cell(row, key):
        j = col.get(key)
        return row[j] if j is not None and j < len(row) else None

    for row in rows[header_at + 1:]:
        if not row or row[0] is None:
            continue
        tip = str(row[0]).strip()
        if tip == "Tip Indicator":  # repeated header mid-sheet
            continue
        value = cell(row, "value")
        if tip.startswith("TOTAL"):
            key = "venituri" if "VENITURI" in tip.upper() else "cheltuieli"
            if value is not None:
                rep.printed_totals[key] = float(value) / LEI_PER_MIE
            continue
        if tip.startswith("FXB"):  # trailing report-id row
            continue
        raw_func = cell(row, "func")
        if value is None or raw_func is None:
            continue

        source = cell(row, "source")
        letter = str(source or "")[:1].upper()
        budget = SOURCE_BUDGET.get(letter)
        if budget is None:
            rep.issues.append(f"sursă de finanțare necunoscută: {source!r}")
            continue
        kind = "revenue" if tip.lower().startswith("venit") else "expense_functional"
        code = dotted_code(raw_func, kind, budget)
        if code is None:
            continue
        raw_econ = cell(row, "econ")
        econ = dotted_code(raw_econ, "expense_economic", budget) if raw_econ else None
        known = True
        if registry is not None:
            known = registry.get(kind, code) is not None
        rep.lines.append(ExecutionLine(
            kind=kind, code=code, econ_code=econ,
            name=str(cell(row, "func_desc") or "").strip()[:90],
            section=SECTION.get(str(cell(row, "section") or "")[:1].upper()),
            budget=budget, source_class=letter,
            value=float(value) / LEI_PER_MIE,
            known_code=known,
        ))

    # the report's own totals are the control sum: they cover every funding
    # source, so compare against all lines regardless of budget
    for key, kind in (("venituri", "revenue"), ("cheltuieli", "expense_functional")):
        printed = rep.printed_totals.get(key)
        if printed is None:
            continue
        got = sum(ln.value for ln in rep.lines if ln.kind == kind)
        if printed and abs(got - printed) / printed > 0.001:
            rep.issues.append(
                f"suma liniilor {key} ({got:.1f}) diferă de totalul tipărit ({printed:.1f})"
            )
    return rep


def capitole(rep: ExecutionReport, kind: str = "expense_functional",
             budget: str = "local") -> list[dict]:
    """Execution rolled up to capitol (``65.02``), descending."""
    agg: dict[str, dict] = {}
    for ln in rep.lines:
        if ln.kind != kind or ln.budget != budget:
            continue
        cap = ".".join(ln.code.split(".")[:2])
        e = agg.setdefault(cap, {"cod": cap, "nume": "", "val": 0.0})
        e["val"] += ln.value
        if not e["nume"] and ln.code == cap:
            e["nume"] = ln.name
    return sorted(agg.values(), key=lambda c: -c["val"])


def _quarter_block(rep: ExecutionReport, quarter: int,
                   registry: Registry | None = None) -> dict:
    """One quarter in full: totals, capitole and its own validation state."""
    caps = capitole(rep, budget="local")
    if registry is not None:
        for c in caps:
            if not c["nume"]:
                e = registry.get("expense_functional", c["cod"])
                if e:
                    c["nume"] = e.name
    return {
        "trimestru": quarter,
        "la_data": rep.report_date,
        "venituri": rep.total("revenue"),
        "cheltuieli": rep.total("expense_functional"),
        "capitole": caps,
        "linii": len(rep.lines),
        "coduri_neenumerate": sum(1 for ln in rep.lines if not ln.known_code),
        "probleme": rep.issues,
    }


def snapshot(reports: dict[int, ExecutionReport], registry: Registry | None = None) -> dict:
    """Per-city execution block: every reported quarter, in full.

    Each quarter is cumulative from 1 January, so the site can show one view
    per quarter. Only the local budget (sources A–D) is reported, to match
    what the site shows for the approved budget.
    """
    if not reports:
        return {}
    quarters = sorted(reports)
    latest_q = quarters[-1]
    blocks = [_quarter_block(reports[q], q, registry) for q in quarters]
    newest = blocks[-1]
    return {
        "unitate": "mii lei",
        "sursa": reports[latest_q].report_type,
        "cif": reports[latest_q].entity_cif,
        "trimestru": latest_q,  # newest reported quarter
        "la_data": newest["la_data"],
        "venituri": newest["venituri"],
        "cheltuieli": newest["cheltuieli"],
        "capitole": newest["capitole"],
        "trimestre": blocks,
        "linii": newest["linii"],
        "coduri_neenumerate": newest["coduri_neenumerate"],
        "probleme": newest["probleme"],
        "note": "Execuție cumulată de la 1 ianuarie, bugetul local (surse A–D). "
                "Raport oficial Forexebug; vezi DISCLAIMER.md.",
    }


# Forexebug publishes a quarter's reports weeks after it closes; this is how
# long we wait before considering a quarter overdue (Q2 2026, closed 30 June,
# was on the portal by late August)
PUBLISH_LAG_DAYS = 55

QUARTER_END = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}

# statuses that count as a verified file — the first three are written by
# data/execution/<year>/download.py, the last by ingest_quarter
VERIFIED_STATUSES = frozenset({
    "copied_and_verified", "downloaded_and_verified", "verified_existing", "verified",
})


def expected_quarter(today: date, year: int) -> int | None:
    """Newest quarter of `year` whose reports should be published by `today`."""
    due = None
    for q, (month, day) in QUARTER_END.items():
        if today >= date(year, month, day) + timedelta(days=PUBLISH_LAG_DAYS):
            due = q
    return due


def quarter_status(exec_root: Path, year: int, today: date) -> dict:
    """What is on disk for a corpus year versus what should be available.

    Drives the scheduled refresh: a quarter is actionable only once its
    manifest exists, because the report URLs come from a CAPTCHA-protected
    portal search and cannot be derived (see docs/executie.md).
    """
    present = sorted(
        q for q in QUARTER_END
        if (exec_root / f"q{q}" / "manifest.json").exists()
    )
    complete = []
    for q in present:
        entries = json.loads((exec_root / f"q{q}" / "manifest.json").read_text()).get("entries", [])
        urls = [e for e in entries if e.get("source_url") or e.get("copy_from")]
        files = [e for e in entries if (exec_root / e["path"]).exists()]
        if entries and len(urls) == len(entries) and len(files) == len(entries):
            complete.append(q)
    due = expected_quarter(today, year)
    missing = [q for q in range(1, (due or 0) + 1) if q not in complete]
    return {
        "an": year,
        "azi": today.isoformat(),
        "trimestre_prezente": present,
        "trimestre_complete": complete,
        "trimestru_asteptat": due,
        "de_adus": missing,
        "urmatorul_de_adus": missing[0] if missing else None,
        "manifest_lipsa": [q for q in missing if q not in present],
    }


def discover_reports(exec_root: Path, county_slug: str, city_slug: str) -> dict[int, Path]:
    """``{1: .../q1/forexebug_execution.xlsx, ...}`` for one city-year."""
    city_dir = exec_root / county_slug / city_slug
    out: dict[int, Path] = {}
    if not city_dir.is_dir():
        return out
    for qdir in sorted(city_dir.glob("q[1-4]")):
        f = qdir / "forexebug_execution.xlsx"
        if f.exists():
            out[int(qdir.name[1:])] = f
    return out


def build_city(exec_root: Path, county_slug: str, city_slug: str,
               registry: Registry | None = None) -> dict:
    """Parse every available quarter for one city and return its snapshot."""
    found = discover_reports(exec_root, county_slug, city_slug)
    reports = {q: parse_report(p, quarter=q, registry=registry) for q, p in found.items()}
    return snapshot(reports, registry)


def scaffold_quarter(exec_root: Path, quarter: int, from_quarter: int | None = None) -> Path:
    """Prepare q<N>/manifest.json from the previous quarter, URLs left empty.

    Everything except the report URL and date is stable between quarters
    (entities, CIFs, paths), so the only manual step left is pasting the
    links returned by the portal search.
    """
    if from_quarter is None:
        existing = [q for q in QUARTER_END
                    if q < quarter and (exec_root / f"q{q}" / "manifest.json").exists()]
        if not existing:
            raise FileNotFoundError(f"niciun trimestru anterior sub {exec_root}")
        from_quarter = max(existing)

    src = json.loads((exec_root / f"q{from_quarter}" / "manifest.json").read_text())
    year = int(src["year"])
    month, day = QUARTER_END[quarter]
    report_date = date(year, month, day).isoformat()

    entries = []
    for e in src.get("entries", []):
        new = dict(e)
        new["path"] = e["path"].replace(f"/q{from_quarter}/", f"/q{quarter}/")
        new["reporting_period"] = f"{year}-Q{quarter}"
        new["report_date"] = report_date
        if "copy_from" in new:
            new["copy_from"] = new["copy_from"].replace(f"/q{from_quarter}/", f"/q{quarter}/")
        else:
            new["source_url"] = None  # to be filled from the portal search
        entries.append(new)

    out = exec_root / f"q{quarter}" / "manifest.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps({
        **{k: v for k, v in src.items() if k != "entries"},
        "quarter": quarter,
        "report_date": report_date,
        "source_audited_on": None,
        "entries": entries,
    }, ensure_ascii=False, indent=1))
    return out


def _sha256(path: Path) -> tuple[str, int]:
    import hashlib

    h, size = hashlib.sha256(), 0
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
            size += len(chunk)
    return h.hexdigest(), size


def ingest_quarter(exec_root: Path, quarter: int,
                   registry: Registry | None = None) -> dict:
    """Take the workbooks placed on disk and make them a verified quarter.

    Files are dropped in by hand after the portal search, so nothing about
    them is assumed: each is parsed, checked against the identity the corpus
    already knows (CIF and entity name from an earlier quarter) and against
    the quarter's own reporting date, then recorded with its checksum.
    A file that fails any check is reported, never silently accepted.
    """
    month, day = QUARTER_END[quarter]
    qdir = exec_root / f"q{quarter}"
    manifest_path = qdir / "manifest.json"

    known = json.loads(manifest_path.read_text()) if manifest_path.exists() else None
    prior_path = qdir / "verification.json"
    prior = {}
    if prior_path.exists():
        prior = {e["path"]: e for e in json.loads(prior_path.read_text()).get("entries", [])
                 if e.get("sha256")}
    template = known or _previous_manifest(exec_root, quarter)
    if template is None:
        raise FileNotFoundError(
            f"niciun manifest de referință sub {exec_root} — nu știu ce entități aștept"
        )
    year = int(template["year"])
    report_date = date(year, month, day).isoformat()
    by_path = {e["path"]: e for e in template.get("entries", [])}

    entries, results, fresh = [], [], False
    for old_path, ref in by_path.items():
        rel = re.sub(r"/q\d+/", f"/q{quarter}/", old_path)
        f = exec_root / rel
        # start from the reference entry so key order — and any field this
        # code does not know about — survives; re-ingesting a good quarter
        # must leave the manifest byte-identical
        entry = dict(ref)
        entry.update({"path": rel, "reporting_period": f"{year}-Q{quarter}",
                      "report_date": report_date})
        if "copy_from" in ref:
            entry["copy_from"] = re.sub(r"/q\d+/", f"/q{quarter}/", ref["copy_from"])
        if known is None:
            entry.pop("source_url", None)  # URLs belong to the quarter they came from

        res = dict(entry)
        if not f.exists():
            res["verification_status"] = "missing"
            fresh = True
            results.append(res)
            entries.append(entry)
            continue

        problems = []
        rep = parse_report(f, quarter=quarter, registry=registry)
        if rep.entity_cif and ref.get("entity_cif") and rep.entity_cif != ref["entity_cif"]:
            problems.append(f"CIF {rep.entity_cif} ≠ {ref['entity_cif']} (așteptat)")
        if rep.report_date and rep.report_date != report_date:
            problems.append(f"data raportului {rep.report_date} ≠ {report_date}")
        if not rep.lines:
            problems.append("raport fără linii de execuție")
        problems += rep.issues

        digest, size = _sha256(f)
        was = prior.get(rel)
        if (was and was.get("sha256") == digest and not problems
                and was.get("verification_status") in VERIFIED_STATUSES):
            # unchanged file, already verified: keep the earlier record intact
            # (it may carry provenance this pass cannot reconstruct)
            results.append(was)
            entries.append(entry)
            continue
        fresh = True  # something was actually (re)verified this pass

        res.update({
            "verification_status": "failed" if problems else "verified",
            "bytes": size, "sha256": digest,
            "lines": len(rep.lines),
            "entity_name_in_file": rep.entity_name,
        })
        if problems:
            res["problems"] = problems
        if not entry.get("source_url"):
            res["provenance"] = "placed_manually"
            entry["provenance"] = "placed_manually"
        results.append(res)
        entries.append(entry)

    ok = [r for r in results if r["verification_status"] in VERIFIED_STATUSES]
    bad = [r for r in results if r["verification_status"] == "failed"]
    missing = [r for r in results if r["verification_status"] == "missing"]

    qdir.mkdir(parents=True, exist_ok=True)
    manifest = dict(template)  # keep the document's own key order
    manifest.update({"quarter": quarter, "report_date": report_date, "entries": entries})
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n")
    if not fresh and prior_path.exists():
        # nothing was re-verified: leave the existing record and checksums alone
        return {"quarter": quarter, "verified": len(ok), "failed": len(bad),
                "missing": len(missing), "entries": results}
    (qdir / "verification.json").write_text(json.dumps({
        "schema_version": 1, "year": year, "quarter": quarter,
        "report_date": report_date, "report_type": "FXB-EXB-901",
        "summary": {"verified": len(ok), "failed": len(bad), "missing": len(missing)},
        "entries": results,
    }, ensure_ascii=False, indent=2) + "\n")
    (qdir / "checksums.sha256").write_text("".join(
        f"{r['sha256']}  {r['path']}\n" for r in results if r.get("sha256")
    ))
    return {"quarter": quarter, "verified": len(ok), "failed": len(bad),
            "missing": len(missing), "entries": results}


def _previous_manifest(exec_root: Path, quarter: int) -> dict | None:
    for q in sorted((q for q in QUARTER_END if q != quarter), reverse=True):
        p = exec_root / f"q{q}" / "manifest.json"
        if p.exists():
            return json.loads(p.read_text())
    return None


def quarters_on_disk(exec_root: Path) -> list[int]:
    """Quarters that have at least one workbook placed under them."""
    found = set()
    for f in exec_root.glob("*/*/q[1-4]/forexebug_execution.xlsx"):
        found.add(int(f.parent.name[1:]))
    return sorted(found)


def write_snapshot(data: dict, out: Path) -> Path:
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    return out
