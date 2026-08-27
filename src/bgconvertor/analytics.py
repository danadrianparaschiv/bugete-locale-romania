"""Auditable city-year analytics derived from the trusted public corpus.

The aggregate keeps extracted budget facts and augmentation side by side but
distinguishable.  This module adds a third layer: reproducible derived metrics.
Every comparison carries an eligibility flag and exclusion reason; missing or
suspect plan totals are never silently ranked as zero.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pydantic import BaseModel, Field

from .aggregate import City, CityYear, Corpus, InflationObservation, build_aggregate

SCHEMA_VERSION = 2
MII_LEI_TO_LEI = 1000.0


class AnalyticsRow(BaseModel):
    year: int
    siruta: str
    municipality: str
    county: str
    county_code: str
    status: str
    artifact_status: str | None = None
    strict_line_rate: float | None = None
    recall_measured: bool | None = None
    plan_comparison_eligible: bool = False
    plan_exclusion_reason: str | None = None
    execution_comparison_eligible: bool = False
    execution_exclusion_reason: str | None = None
    population: int | None = None
    population_date: str | None = None
    area_km2: float | None = None
    density_per_km2: float | None = None
    planned_revenue_mii_lei: float | None = None
    planned_expense_mii_lei: float | None = None
    planned_balance_mii_lei: float | None = None
    planned_revenue_lei_per_capita: float | None = None
    planned_expense_lei_per_capita: float | None = None
    planned_balance_lei_per_capita: float | None = None
    plan_expense_per_capita_rank: int | None = None
    plan_expense_rank_cohort: int | None = None
    execution_quarter: int | None = None
    execution_date: str | None = None
    actual_revenue_mii_lei: float | None = None
    actual_expense_mii_lei: float | None = None
    actual_balance_mii_lei: float | None = None
    actual_revenue_lei_per_capita: float | None = None
    actual_expense_lei_per_capita: float | None = None
    actual_balance_lei_per_capita: float | None = None
    execution_expense_per_capita_rank: int | None = None
    execution_expense_rank_cohort: int | None = None
    revenue_execution_pct: float | None = None
    expense_execution_pct: float | None = None
    planned_revenue_yoy_pct: float | None = None
    planned_expense_yoy_pct: float | None = None
    regional_classification_version: str | None = None
    nuts1_code: str | None = None
    nuts1_name: str | None = None
    nuts2_code: str | None = None
    nuts2_name: str | None = None
    nuts3_code: str | None = None
    nuts3_name: str | None = None
    hicp_annual_average_rate_pct: float | None = None
    inflation_status: str = "full_year_not_available"
    planned_revenue_yoy_real_pct: float | None = None
    planned_expense_yoy_real_pct: float | None = None


class ChapterRow(BaseModel):
    year: int
    siruta: str
    municipality: str
    county: str
    code: str
    name: str
    total_mii_lei: float
    share_of_planned_expense_pct: float | None = None
    lei_per_capita: float | None = None
    note: str = "Capitol prezent în topul verificat al municipiului; lista nu este exhaustivă."


class AnalyticsDataset(BaseModel):
    schema_version: int = SCHEMA_VERSION
    unit: str = "mii lei; indicatorii per capita sunt lei/locuitor"
    comparison_policy: str = (
        "Clasamentele includ numai bundle-uri publice coerente, cu total de cheltuieli "
        "verificabil, populație RPL2021 și fără semnal de plan incomplet."
    )
    sources: dict = Field(default_factory=dict)
    coverage: dict = Field(default_factory=dict)
    rows: list[AnalyticsRow] = Field(default_factory=list)
    chapters: list[ChapterRow] = Field(default_factory=list)

    def year_rows(self, year: int) -> list[AnalyticsRow]:
        return [row for row in self.rows if row.year == year]


def _strict_rate(cy: CityYear) -> float | None:
    if not cy.quality:
        return None
    if cy.quality.pct_lines_strictly_verified is not None:
        return cy.quality.pct_lines_strictly_verified
    return cy.quality.pct_clean


def _plan_exclusion(city: City, cy: CityYear) -> str | None:
    if cy.status != "converted":
        return f"status_{cy.status}"
    if not cy.has_analysis:
        return "analysis_lipsa"
    if not cy.totals_mii_lei.get("cheltuieli"):
        return "total_cheltuieli_lipsa"
    if cy.executie is not None and cy.executie.plan_incomplet:
        return "plan_incomplet_fata_de_executie"
    if not city.populatie:
        return "populatie_lipsa"
    return None


def _execution_exclusion(city: City, cy: CityYear) -> str | None:
    if cy.executie is None:
        return "executie_lipsa"
    if cy.executie.cheltuieli is None:
        return "total_executie_cheltuieli_lipsa"
    if not city.populatie:
        return "populatie_lipsa"
    return None


def _pct_change(previous: float | None, current: float | None) -> float | None:
    if previous in (None, 0) or current is None:
        return None
    return round((current - previous) / previous * 100, 2)


def _real_change(nominal_change: float | None, inflation: float | None) -> float | None:
    """Deflate a nominal year-over-year change by observed annual HICP."""
    if nominal_change is None or inflation is None:
        return None
    return round(((1 + nominal_change / 100) / (1 + inflation / 100) - 1) * 100, 2)


def _rank(
    rows: list[AnalyticsRow], value_field: str, rank_field: str, cohort_field: str,
    eligible_field: str,
) -> None:
    candidates = [
        row for row in rows
        if getattr(row, eligible_field) and getattr(row, value_field) is not None
    ]
    values = [getattr(row, value_field) for row in candidates]
    for row in candidates:
        value = getattr(row, value_field)
        setattr(row, rank_field, 1 + sum(other > value for other in values))
        setattr(row, cohort_field, len(candidates))


def _row(
    city: City,
    year: int,
    cy: CityYear,
    inflation: InflationObservation | None,
) -> AnalyticsRow:
    plan_reason = _plan_exclusion(city, cy)
    execution_reason = _execution_exclusion(city, cy)
    population = city.populatie
    area = city.suprafata_km2
    planned_revenue = cy.totals_mii_lei.get("venituri")
    planned_expense = cy.totals_mii_lei.get("cheltuieli")
    plan_eligible = plan_reason is None
    execution_eligible = execution_reason is None
    execution = cy.executie
    region = city.regional_classification

    row = AnalyticsRow(
        year=year,
        siruta=city.siruta,
        municipality=city.name,
        county=city.county,
        county_code=city.county_code,
        status=cy.status,
        artifact_status=cy.artifact_status,
        strict_line_rate=_strict_rate(cy),
        recall_measured=cy.quality.recall_measured if cy.quality else None,
        plan_comparison_eligible=plan_eligible,
        plan_exclusion_reason=plan_reason,
        execution_comparison_eligible=execution_eligible,
        execution_exclusion_reason=execution_reason,
        population=population,
        population_date=city.populatie_data,
        area_km2=area,
        density_per_km2=round(population / area, 2) if population and area else None,
        planned_revenue_mii_lei=planned_revenue,
        planned_expense_mii_lei=planned_expense,
        execution_quarter=execution.trimestru if execution else None,
        execution_date=execution.la_data if execution else None,
        actual_revenue_mii_lei=execution.venituri if execution else None,
        actual_expense_mii_lei=execution.cheltuieli if execution else None,
        revenue_execution_pct=(
            execution.pct_venituri if execution and plan_eligible else None
        ),
        expense_execution_pct=(
            execution.pct_cheltuieli if execution and plan_eligible else None
        ),
        regional_classification_version=(region.dataset_version if region else None),
        nuts1_code=region.nuts1_code if region else None,
        nuts1_name=region.nuts1_name if region else None,
        nuts2_code=region.nuts2_code if region else None,
        nuts2_name=region.nuts2_name if region else None,
        nuts3_code=region.nuts3_code if region else None,
        nuts3_name=region.nuts3_name if region else None,
        hicp_annual_average_rate_pct=(
            inflation.annual_average_rate_pct if inflation else None
        ),
        inflation_status=inflation.status if inflation else "full_year_not_available",
    )
    if plan_eligible and planned_revenue is not None and planned_expense is not None:
        row.planned_balance_mii_lei = round(planned_revenue - planned_expense, 3)
    if plan_eligible and population:
        if planned_revenue is not None:
            row.planned_revenue_lei_per_capita = round(
                planned_revenue * MII_LEI_TO_LEI / population, 2
            )
        row.planned_expense_lei_per_capita = round(
            planned_expense * MII_LEI_TO_LEI / population, 2
        )
        if row.planned_balance_mii_lei is not None:
            row.planned_balance_lei_per_capita = round(
                row.planned_balance_mii_lei * MII_LEI_TO_LEI / population, 2
            )
    if execution and execution.venituri is not None and execution.cheltuieli is not None:
        row.actual_balance_mii_lei = round(execution.venituri - execution.cheltuieli, 3)
    if execution_eligible and population:
        if execution and execution.venituri is not None:
            row.actual_revenue_lei_per_capita = round(
                execution.venituri * MII_LEI_TO_LEI / population, 2
            )
        row.actual_expense_lei_per_capita = round(
            execution.cheltuieli * MII_LEI_TO_LEI / population, 2
        )
        if row.actual_balance_mii_lei is not None:
            row.actual_balance_lei_per_capita = round(
                row.actual_balance_mii_lei * MII_LEI_TO_LEI / population, 2
            )
    return row


def build_analytics(corpus: Corpus) -> AnalyticsDataset:
    rows = [
        _row(city, int(year), cy, corpus.inflation.get(str(year)))
        for city in corpus.cities
        for year, cy in city.years.items()
    ]

    by_city: dict[str, list[AnalyticsRow]] = {}
    for row in rows:
        by_city.setdefault(row.siruta, []).append(row)
    for city_rows in by_city.values():
        previous: AnalyticsRow | None = None
        for row in sorted(city_rows, key=lambda item: item.year):
            if row.plan_comparison_eligible:
                if previous is not None:
                    row.planned_revenue_yoy_pct = _pct_change(
                        previous.planned_revenue_mii_lei, row.planned_revenue_mii_lei
                    )
                    row.planned_expense_yoy_pct = _pct_change(
                        previous.planned_expense_mii_lei, row.planned_expense_mii_lei
                    )
                    if row.year == previous.year + 1:
                        row.planned_revenue_yoy_real_pct = _real_change(
                            row.planned_revenue_yoy_pct,
                            row.hicp_annual_average_rate_pct,
                        )
                        row.planned_expense_yoy_real_pct = _real_change(
                            row.planned_expense_yoy_pct,
                            row.hicp_annual_average_rate_pct,
                        )
                previous = row

    for year in corpus.years:
        year_rows = [row for row in rows if row.year == year]
        execution_periods: dict[tuple[int | None, str | None], int] = {}
        for row in year_rows:
            if row.execution_comparison_eligible:
                period = (row.execution_quarter, row.execution_date)
                execution_periods[period] = execution_periods.get(period, 0) + 1
        if execution_periods:
            comparable_period = max(
                execution_periods,
                key=lambda period: (
                    execution_periods[period], period[1] or "", period[0] or 0,
                ),
            )
            for row in year_rows:
                if (
                    row.execution_comparison_eligible
                    and (row.execution_quarter, row.execution_date) != comparable_period
                ):
                    row.execution_comparison_eligible = False
                    row.execution_exclusion_reason = "perioada_executie_necomparabila"
        _rank(
            year_rows, "planned_expense_lei_per_capita",
            "plan_expense_per_capita_rank", "plan_expense_rank_cohort",
            "plan_comparison_eligible",
        )
        _rank(
            year_rows, "actual_expense_lei_per_capita",
            "execution_expense_per_capita_rank", "execution_expense_rank_cohort",
            "execution_comparison_eligible",
        )

    row_by_key = {(row.siruta, row.year): row for row in rows}
    chapters = []
    for city in corpus.cities:
        for year_text, cy in city.years.items():
            year = int(year_text)
            analytics_row = row_by_key[(city.siruta, year)]
            if not analytics_row.plan_comparison_eligible:
                continue
            planned = analytics_row.planned_expense_mii_lei
            for chapter in cy.top_capitole:
                chapters.append(ChapterRow(
                    year=year,
                    siruta=city.siruta,
                    municipality=city.name,
                    county=city.county,
                    code=chapter.code,
                    name=chapter.name,
                    total_mii_lei=chapter.total,
                    share_of_planned_expense_pct=(
                        round(chapter.total / planned * 100, 2) if planned else None
                    ),
                    lei_per_capita=(
                        round(chapter.total * MII_LEI_TO_LEI / city.populatie, 2)
                        if city.populatie else None
                    ),
                ))

    coverage = {}
    for year in corpus.years:
        year_rows = [row for row in rows if row.year == year]
        coverage[str(year)] = {
            "municipalities": len(year_rows),
            "trusted_plan_analyses": sum(row.status == "converted" for row in year_rows),
            "plan_comparison_eligible": sum(
                row.plan_comparison_eligible for row in year_rows
            ),
            "execution_comparison_eligible": sum(
                row.execution_comparison_eligible for row in year_rows
            ),
            "recall_measured": sum(row.recall_measured is True for row in year_rows),
        }
    sources = {
        "budget_plan": {
            "description": "analysis.json din bundle-ul public auditat al fiecărui PDF bugetar",
            "join_key": "SIRUTA + an",
        },
        "execution": {
            "description": "rapoarte trimestriale Forexebug, păstrate separat de plan",
            "join_key": "SIRUTA + an",
        },
        **corpus.augmentation_sources,
    }
    return AnalyticsDataset(
        sources=sources,
        coverage=coverage,
        rows=sorted(rows, key=lambda row: (-row.year, row.county_code, row.municipality)),
        chapters=sorted(
            chapters, key=lambda row: (-row.year, row.county, row.municipality, -row.total_mii_lei)
        ),
    )


def build_from_data(data_root: Path) -> AnalyticsDataset:
    return build_analytics(build_aggregate(data_root))


def _atomic_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, delete=False) as handle:
        handle.write(content)
        temporary = Path(handle.name)
    temporary.replace(path)


def write_json(dataset: AnalyticsDataset, path: Path) -> Path:
    _atomic_text(path, dataset.model_dump_json(indent=1, exclude_none=True) + "\n")
    return path


def write_csv(dataset: AnalyticsDataset, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(AnalyticsRow.model_fields)
    with NamedTemporaryFile("w", encoding="utf-8", newline="", dir=path.parent, delete=False) as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in dataset.rows:
            writer.writerow(row.model_dump())
        temporary = Path(handle.name)
    temporary.replace(path)
    return path


def _style_sheet(ws, widths: dict[int, float]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def _add_table(ws, name: str) -> None:
    if ws.max_row < 2:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def write_workbook(dataset: AnalyticsDataset, path: Path) -> Path:
    """Write an analysis-ready workbook with formulas for derived columns."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Sumar"
    summary.sheet_view.showGridLines = False
    summary["A1"] = "Analitice bugete locale"
    summary["A1"].font = Font(size=18, bold=True, color="1F4E78")
    summary.merge_cells("A1:D1")
    summary["A3"] = "Contract"
    summary["B3"] = dataset.comparison_policy
    summary["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["A4"] = "Unități"
    summary["B4"] = dataset.unit
    summary["A6"] = "An"
    summary["B6"] = "Municipii"
    summary["C6"] = "Eligibile plan"
    summary["D6"] = "Eligibile execuție"
    for cell in summary[6]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for values in sorted(dataset.coverage.items(), reverse=True):
        year, metrics = values
        summary.append([
            int(year), metrics["municipalities"], metrics["plan_comparison_eligible"],
            metrics["execution_comparison_eligible"],
        ])
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 68
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 20
    summary.row_dimensions[3].height = 60

    ws = wb.create_sheet("Municipii")
    headers = [
        "An", "SIRUTA", "Municipiu", "Județ", "Eligibil plan", "Motiv excludere plan",
        "Bundle", "% linii strict", "Recall măsurat", "Populație", "Data populației",
        "Suprafață km²", "Densitate loc/km²", "Venituri plan mii lei",
        "Cheltuieli plan mii lei", "Sold plan mii lei", "Venituri lei/loc",
        "Cheltuieli lei/loc", "Sold plan lei/loc", "Rang cheltuieli plan/loc",
        "Cohortă rang plan", "Trimestru execuție", "Data execuției",
        "Venituri efective mii lei", "Cheltuieli efective mii lei",
        "Sold efectiv mii lei", "Venituri efective lei/loc",
        "Cheltuieli efective lei/loc", "Sold efectiv lei/loc",
        "Rang cheltuieli efective/loc", "Cohortă rang execuție", "% execuție venituri",
        "% execuție cheltuieli", "Δ venituri plan %", "Δ cheltuieli plan %",
        "Eligibil execuție", "Motiv excludere execuție",
        "Versiune clasificare regională", "NUTS1 cod", "NUTS1 nume",
        "NUTS2 cod", "NUTS2 nume", "NUTS3 cod", "NUTS3 nume",
        "Inflație HICP medie anuală %", "Stare inflație",
        "Δ real venituri plan %", "Δ real cheltuieli plan %",
    ]
    ws.append(headers)
    for row_number, row in enumerate(dataset.rows, start=2):
        ws.append([
            row.year, row.siruta, row.municipality, row.county,
            "da" if row.plan_comparison_eligible else "nu",
            row.plan_exclusion_reason, row.artifact_status,
            row.strict_line_rate,
            "da" if row.recall_measured is True else "nu" if row.recall_measured is False else None,
            row.population, row.population_date,
            row.area_km2,
            f'=IFERROR(J{row_number}/L{row_number},"")',
            row.planned_revenue_mii_lei, row.planned_expense_mii_lei,
            f'=IF(OR($E{row_number}<>"da",N{row_number}="",O{row_number}=""),"",N{row_number}-O{row_number})',
            f'=IF(OR($E{row_number}<>"da",N{row_number}=""),"",IFERROR(N{row_number}*1000/J{row_number},""))',
            f'=IF(OR($E{row_number}<>"da",O{row_number}=""),"",IFERROR(O{row_number}*1000/J{row_number},""))',
            f'=IF(OR($E{row_number}<>"da",P{row_number}=""),"",IFERROR(P{row_number}*1000/J{row_number},""))',
            row.plan_expense_per_capita_rank, row.plan_expense_rank_cohort,
            row.execution_quarter, row.execution_date, row.actual_revenue_mii_lei,
            row.actual_expense_mii_lei,
            f'=IF(OR(X{row_number}="",Y{row_number}=""),"",X{row_number}-Y{row_number})',
            f'=IF(OR($AJ{row_number}<>"da",X{row_number}=""),"",IFERROR(X{row_number}*1000/J{row_number},""))',
            f'=IF(OR($AJ{row_number}<>"da",Y{row_number}=""),"",IFERROR(Y{row_number}*1000/J{row_number},""))',
            f'=IF(OR($AJ{row_number}<>"da",Z{row_number}=""),"",IFERROR(Z{row_number}*1000/J{row_number},""))',
            row.execution_expense_per_capita_rank, row.execution_expense_rank_cohort,
            row.revenue_execution_pct, row.expense_execution_pct,
            row.planned_revenue_yoy_pct, row.planned_expense_yoy_pct,
            "da" if row.execution_comparison_eligible else "nu",
            row.execution_exclusion_reason,
            row.regional_classification_version,
            row.nuts1_code, row.nuts1_name, row.nuts2_code, row.nuts2_name,
            row.nuts3_code, row.nuts3_name,
            row.hicp_annual_average_rate_pct, row.inflation_status,
            row.planned_revenue_yoy_real_pct, row.planned_expense_yoy_real_pct,
        ])
    _style_sheet(ws, {
        1: 8, 2: 12, 3: 24, 4: 20, 5: 14, 6: 32, 7: 16, 8: 14, 9: 14,
        10: 14, 11: 16, 12: 14, 13: 16, 14: 19, 15: 20, 16: 18, 17: 18,
        18: 20, 19: 18, 20: 20, 21: 16, 22: 16, 23: 16, 24: 20, 25: 21,
        26: 18, 27: 21, 28: 22, 29: 19, 30: 23, 31: 18, 32: 19, 33: 20,
        34: 18, 35: 20, 36: 17, 37: 34, 38: 22, 39: 12, 40: 24,
        41: 12, 42: 24, 43: 12, 44: 24, 45: 21, 46: 26, 47: 20, 48: 22,
    })
    _add_table(ws, "MunicipiiAnalytics")
    formats = {
        8: "0.0", 10: "#,##0", 12: "#,##0.00", 13: "#,##0.00",
        **{column: "#,##0.00" for column in range(14, 20)},
        20: "0", 21: "0", 22: "0",
        **{column: "#,##0.00" for column in range(24, 30)},
        30: "0", 31: "0",
        **{column: "0.00" for column in range(32, 36)},
        45: "0.00", 47: "0.00", 48: "0.00",
    }
    for column, number_format in formats.items():
        for cells in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for cell in cells:
                cell.number_format = number_format

    chapters = wb.create_sheet("Capitole")
    chapters.append([
        "An", "SIRUTA", "Municipiu", "Județ", "Cod", "Capitol", "Total mii lei",
        "% din cheltuieli plan", "Lei/locuitor", "Notă",
    ])
    for chapter in dataset.chapters:
        chapters.append([
            chapter.year, chapter.siruta, chapter.municipality, chapter.county,
            chapter.code, chapter.name, chapter.total_mii_lei,
            chapter.share_of_planned_expense_pct, chapter.lei_per_capita, chapter.note,
        ])
    _style_sheet(chapters, {1: 8, 2: 12, 3: 24, 4: 20, 5: 12, 6: 42, 10: 62})
    _add_table(chapters, "CapitoleAnalytics")

    sources = wb.create_sheet("Surse augmentare")
    sources.append(["Strat", "Câmp", "Valoare"])
    for layer, metadata in dataset.sources.items():
        if isinstance(metadata, dict):
            for key, value in metadata.items():
                cell_value = (
                    json.dumps(value, ensure_ascii=False, sort_keys=True)
                    if isinstance(value, (dict, list)) else value
                )
                sources.append([layer, key, cell_value])
        else:
            sources.append([layer, "descriere", metadata])
    _style_sheet(sources, {1: 22, 2: 24, 3: 92})
    _add_table(sources, "SurseAnalytics")
    for row in sources.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    with NamedTemporaryFile(suffix=".xlsx", dir=path.parent, delete=False) as handle:
        temporary = Path(handle.name)
    wb.save(temporary)
    temporary.replace(path)
    return path


def write_outputs(dataset: AnalyticsDataset, out_dir: Path) -> dict[str, Path]:
    return {
        "json": write_json(dataset, out_dir / "analytics.json"),
        "csv": write_csv(dataset, out_dir / "analytics.csv"),
        "xlsx": write_workbook(dataset, out_dir / "analytics.xlsx"),
    }
