"""Excel export: one workbook per PDF.

Sheets: data per document x section (venituri + cheltuieli), a 'Probleme'
sheet listing every Issue, and a 'Sumar calitate' scorecard.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .model import BudgetDocument, ConversionResult

# Canonical order + labels; each sheet shows only the columns its lines carry.
COLUMN_LABELS = [
    ("total", "TOTAL 2026"),
    ("total_2026", "TOTAL 2026"),
    ("credite_stinse", "Credite stingere plati"),
    ("credite_restante", "Credite plati restante"),
    ("trim1", "Trim. I"),
    ("trim2", "Trim. II"),
    ("trim3", "Trim. III"),
    ("trim4", "Trim. IV"),
    ("est2027", "Estimare 2027"),
    ("est2028", "Estimare 2028"),
    ("est2029", "Estimare 2029"),
    ("buget_2026", "Buget 2026 (initial)"),
    ("influente", "Influente"),
    ("valoare_an_curent", "Valoare an curent"),
    ("buget_local", "Buget local"),
    ("bugetul_local", "Bugetul local"),
    ("inst_venituri_proprii_subventii", "Inst. venituri proprii si subventii"),
    ("inst_integral_venituri_proprii", "Inst. integral venituri proprii"),
    ("imprumuturi", "Imprumuturi"),
    ("imprumuturi_externe", "Imprumuturi externe"),
    ("imprumuturi_interne", "Imprumuturi interne"),
    ("fonduri_externe", "Fonduri externe nerambursabile"),
    ("transferuri", "Transferuri intre bugete"),
    ("total_general", "Total buget general"),
    ("credite_externe", "Credite externe"),
    ("credite_interne", "Credite interne"),
    ("buget_fen", "Buget FEN"),
]
_ORDER = {k: i for i, (k, _) in enumerate(COLUMN_LABELS)}
_LABELS = dict(COLUMN_LABELS)


def _sheet_columns(lines) -> list[tuple[str, str]]:
    present = {c for ln in lines for c in (*ln.values, *ln.x_markers)}
    ordered = sorted(present, key=lambda c: _ORDER.get(c, 99))
    return [(c, _LABELS.get(c, c)) for c in ordered]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
SECTION_FONT = Font(bold=True)

DOC_PREFIX = {"local": "BL", "own_revenue": "VP", "general": "BG", "unknown": "DOC"}


def _append_values(ws, values) -> None:
    """Append literal values without interpreting PDF text as Excel formulas."""
    ws.append(values)
    for cell in ws[ws.max_row]:
        if cell.data_type == "f" and isinstance(cell.value, str):
            cell.data_type = "s"


def export(
    result: ConversionResult,
    out_path: Path,
    publication: dict | None = None,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()
    for doc in result.documents:
        prefix = DOC_PREFIX.get(doc.budget, "DOC")
        if prefix in used_names:
            prefix = f"{prefix}{len(used_names)}"
        used_names.add(prefix)
        canonical = ("TOTAL", "FUNCTIONARE", "DEZVOLTARE")
        emitted = 0
        for section in canonical:
            lines = doc.section_lines(section)
            if lines:
                _data_sheet(wb, f"{prefix} {section[:10]}", doc, lines)
                emitted += len(lines)
        # scanned annexes often carry no canonical section markers — everything
        # else (None or ad-hoc contexts) goes to one Date sheet, in page order
        rest = [ln for ln in doc.lines if ln.section not in canonical]
        if rest:
            _data_sheet(wb, f"{prefix} Date", doc, rest)

    _issues_sheet(wb, result)
    _summary_sheet(wb, result, publication=publication)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{out_path.stem}.", suffix=out_path.suffix or ".xlsx", dir=out_path.parent
    )
    os.close(fd)
    try:
        wb.save(tmp_name)
        os.replace(tmp_name, out_path)
    except BaseException:
        try:
            os.unlink(tmp_name)
        except FileNotFoundError:
            pass
        raise
    return out_path


def _data_sheet(wb: Workbook, name: str, doc: BudgetDocument, lines) -> None:
    ws = wb.create_sheet(name[:31])
    value_columns = _sheet_columns(lines)
    headers = ["Cod", "Cod functional", "Denumire", "Rand", "Pag.", "Tip"] + [
        h for _, h in value_columns
    ] + ["Sursa", "Probleme"]
    _append_values(ws, headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for ln in lines:
        row = [
            ln.code,
            ln.func_code,
            ln.name,
            ln.row_no,
            ln.page,
            ln.kind if ln.kind != "heading" else "",
        ]
        for col, _ in value_columns:
            if col in ln.x_markers:
                row.append("X")
            else:
                v = ln.values.get(col)
                row.append(float(v) if v is not None else None)
        row.append(ln.source)
        row.append("; ".join(i.message for i in ln.issues) or None)
        _append_values(ws, row)

        excel_row = ws.max_row
        if ln.kind == "heading":
            ws.cell(row=excel_row, column=3).font = SECTION_FONT
        severities = {i.severity for i in ln.issues}
        fill = ERROR_FILL if "error" in severities else WARN_FILL if "warning" in severities else None
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=c).fill = fill

    widths = [12, 12, 60, 6, 6, 18] + [14] * len(value_columns) + [8, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=6 + len(value_columns)):
        for cell in row:
            cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"


def _issues_sheet(wb: Workbook, result: ConversionResult) -> None:
    ws = wb.create_sheet("Probleme")
    _append_values(ws, ["Verificare", "Severitate", "Pagina", "Cod", "Coloana", "Mesaj"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for issue in result.all_issues():
        _append_values(
            ws,
            [issue.check, issue.severity, issue.page, issue.code, issue.column, issue.message],
        )
        if issue.severity == "error":
            ws.cell(row=ws.max_row, column=2).fill = ERROR_FILL
    for i, w in enumerate([16, 10, 8, 14, 12, 100], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _summary_sheet(
    wb: Workbook,
    result: ConversionResult,
    publication: dict | None = None,
) -> None:
    ws = wb.create_sheet("Sumar calitate")
    stats = result.stats()
    rows = [
        ("Fisier", result.pdf),
        ("Schema calitate", stats["quality_schema_version"]),
        ("Metrica de calitate", stats["metric"]),
        ("Recall masurat", "da" if stats["recall_measured"] else "nu"),
        ("Documente", stats["documents"]),
        ("Linii de date", stats["lines"]),
        ("Linii strict verificate", stats["lines_strictly_verified"]),
        ("% linii strict verificate", stats["pct_lines_strictly_verified"]),
        ("Celule numerice", stats["numeric_cells"]),
        ("Celule numerice strict verificate", stats["numeric_cells_strictly_verified"]),
        ("% celule numerice strict verificate", stats["pct_numeric_cells_strictly_verified"]),
        ("Pagini asteptate", stats["scope"]["pages_expected"]),
        ("Pagini selectate", stats["scope"]["pages_selected"]),
        ("Pagini procesate", stats["scope"]["pages_processed"]),
        ("PDF complet", "da" if stats["scope"]["complete_pdf"] else "nu"),
        # Legacy labels retained for older readers.
        ("Linii fara probleme", stats["lines_clean"]),
        ("% curat", stats["pct_clean"]),
        ("Erori", stats["issues"]["error"]),
        ("Avertismente", stats["issues"]["warning"]),
        ("Informatii", stats["issues"]["info"]),
    ]
    if publication:
        rows.extend([
            ("Schema publicare", publication.get("schema_version")),
            ("Bundle conversie", publication.get("bundle_id")),
            ("SHA-256 PDF sursa", publication.get("source_pdf_sha256")),
        ])
    llm_models = sorted({
        ln.source.split(":", 1)[1] if ":" in ln.source else "model neînregistrat"
        for doc in result.documents for ln in doc.lines
        if ln.source.startswith("llm")
    })
    if llm_models:
        rows.append(("Modele LLM folosite", ", ".join(llm_models)))
    for doc in result.documents:
        rows.append((f"— {doc.title[:60]}", f"{doc.budget}, pag. {doc.pages[0]}-{doc.pages[-1]}, {len(doc.lines)} linii"))
    for label, value in rows:
        _append_values(ws, [label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60
