"""Corpus aggregate + multi-year site build, on a synthetic data/ tree."""

import json

import pytest
from openpyxl import Workbook

from bgconvertor import aggregate as agg
from bgconvertor.analytics import build_analytics, write_outputs
from bgconvertor.site import build_all


def _quality_workbook(path, lines=100, pct=95.0, errors=0, warnings=5):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sumar calitate"
    for row in (
        ("Linii de date", lines), ("% curat", pct),
        ("Erori", errors), ("Avertismente", warnings),
    ):
        ws.append(row)
    wb.save(path)


def _write_year(data_root, year, converted, totals=(123.4, 120.0)):
    ydir = data_root / str(year)
    city = ydir / "01-alba" / "1017-alba-iulia"
    city.mkdir(parents=True)
    (city / "budget_file.pdf").write_bytes(b"%PDF-fake")
    entry = {
        "county_code": "01", "county_name": "Alba",
        "capital_siruta": "1017", "capital_name": "Alba Iulia",
        "path": "01-alba/1017-alba-iulia/budget_file.pdf",
        "source_url": f"https://example.ro/{year}.pdf",
    }
    if converted:
        entry["conversion"] = {
            "status": "converted", "lines": 100, "pct_clean": 95.0,
            "errors": 0, "warnings": 5,
        }
        entry["timeline"] = {"debate_date": f"{year}-04-01", "approved_date": f"{year}-04-28",
                             "hcl": f"HCL 1/{year}"}
        _quality_workbook(city / "budget_file.xlsx")
        (city / "analysis.json").write_text(json.dumps({
            "quality": {"lines": 100, "pct_clean": 95.0, "errors": 0, "warnings": 5, "documents": 1},
            "totals_mii_lei": {"venituri": totals[0], "cheltuieli": totals[1]},
            "top_capitole": [{"code": "65.02", "name": "Învățământ", "total": 50.0}],
        }))
    (ydir / "manifest.json").write_text(json.dumps({"year": year, "entries": [entry]}))


@pytest.fixture
def data_root(tmp_path):
    root = tmp_path / "data"
    _write_year(root, 2026, converted=True)
    _write_year(root, 2025, converted=False)
    return root


def test_aggregate_merges_years_by_siruta(data_root):
    corpus = agg.build_aggregate(data_root)
    assert corpus.years == [2026, 2025]
    (city,) = corpus.cities
    assert city.siruta == "1017" and city.county_code == "01"
    assert set(city.years) == {"2026", "2025"}

    cy26 = city.years["2026"]
    assert cy26.status == "converted" and cy26.has_analysis
    assert cy26.quality.pct_clean == 95.0
    assert cy26.totals_mii_lei["venituri"] == 123.4
    assert cy26.top_capitole[0].code == "65.02"
    assert cy26.timeline.hcl == "HCL 1/2026"
    assert cy26.files.pdf == "data/2026/01-alba/1017-alba-iulia/budget_file.pdf"
    assert cy26.files.xlsx.endswith(".xlsx")

    cy25 = city.years["2025"]
    assert cy25.status == "pending" and not cy25.has_analysis
    assert cy25.quality is None
    assert cy25.files.pdf and cy25.files.xlsx is None


def test_oversize_pdf_not_linked(data_root, monkeypatch):
    monkeypatch.setattr(agg, "GITHUB_FILE_LIMIT", 1)
    corpus = agg.build_aggregate(data_root)
    cy = corpus.cities[0].years["2026"]
    assert cy.files.pdf is None
    assert cy.files.source_url == "https://example.ro/2026.pdf"


def test_native_excel_source_is_labeled_and_linked(data_root, tmp_path):
    ydir = data_root / "2024"
    city = ydir / "01-alba" / "1017-alba-iulia"
    city.mkdir(parents=True)
    (city / "buget_orig.xlsx").write_bytes(b"native-source")
    (ydir / "manifest.json").write_text(json.dumps({
        "year": 2024,
        "entries": [{
            "county_code": "01", "county_name": "Alba",
            "capital_siruta": "1017", "capital_name": "Alba Iulia",
            "path": "01-alba/1017-alba-iulia/buget_orig.xlsx",
            "source_format": "xlsx",
            "source_url": "https://example.ro/2024.xlsx",
        }],
    }))

    corpus = agg.build_aggregate(data_root)
    cy = corpus.cities[0].years["2024"]
    assert cy.files.pdf is None
    assert cy.files.source == "data/2024/01-alba/1017-alba-iulia/buget_orig.xlsx"
    assert cy.files.source_format == "xlsx"

    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")
    index = (out / "2024" / "index.html").read_text()
    assert "sursă xlsx" in index
    assert "buget_orig.xlsx" in index


def test_uncommitted_raw_pdf_policy_links_to_official_source(data_root, tmp_path):
    manifest_path = data_root / "2026" / "manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["raw_pdf_policy"] = \
        "excluded_from_git_with_committed_urls_checksums_and_derived_bundles"
    manifest_path.write_text(json.dumps(manifest))

    corpus = agg.build_aggregate(data_root)
    cy = corpus.cities[0].years["2026"]
    assert cy.files.pdf is None
    assert cy.files.source_url == "https://example.ro/2026.pdf"

    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")
    index = (out / "index.html").read_text()
    assert 'href="https://example.ro/2026.pdf">sursă</a>' in index
    assert "data/2026/01-alba/1017-alba-iulia/budget_file.pdf" not in index


def test_build_all_writes_years_and_data_endpoint(data_root, tmp_path):
    out = tmp_path / "site"
    results = build_all(data_root, out, base_url="/repo")
    assert [r["converted_pages"] for r in results] == [1, 0]

    index = (out / "index.html").read_text()
    assert "Bugetele locale 2026" in index
    assert '<a href="/repo/2025/">2025</a>' in index
    assert (out / "city" / "1017.html").exists()

    index25 = (out / "2025" / "index.html").read_text()
    assert "Bugetele locale 2025" in index25
    assert '<a href="/repo/">2026</a>' in index25
    assert not (out / "2025" / "city" / "1017.html").exists()

    data = json.loads((out / "data" / "corpus.json").read_text())
    assert data["schema_version"] == agg.SCHEMA_VERSION == 4
    assert data["years"] == [2026, 2025]
    assert data["cities"][0]["years"]["2026"]["totals_mii_lei"]["cheltuieli"] == 120.0

    # the animated procedure explainer, per edition, linked from the index
    proc = (out / "procedura.html").read_text()
    assert "45 de zile" in proc and "28.04.2026" in proc  # stats from newest year
    assert (out / "2025" / "procedura.html").exists()
    assert '/repo/procedura.html' in index
    assert '/repo/2025/procedura.html' in index25

    # a single year with figures -> no year-over-year section
    assert "Evoluție an-la-an" not in (out / "city" / "1017.html").read_text()


def test_every_site_table_has_accessible_sort_controls(tmp_path):
    data_root = tmp_path / "sortable-data"
    _write_year(data_root, 2026, converted=True)
    _write_year(data_root, 2025, converted=True)
    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")

    pages = [
        out / "index.html",
        out / "comparatii.html",
        out / "city" / "1017.html",
    ]
    for path in pages:
        page = path.read_text()
        assert page.count("<table") == page.count('<table class="sortable">')
        assert page.count("<table") == page.count("<thead>")
        assert page.count("<table") == page.count("<tbody>")
        assert 'document.querySelectorAll("table.sortable")' in page
        assert 'header.setAttribute("aria-sort", direction)' in page

    city = (out / "city" / "1017.html").read_text()
    assert "<th>Etapă</th>" in city
    assert "<th>Indicator</th><th>Valoare</th>" in city
    index = (out / "index.html").read_text()
    assert '<th data-sort-type="date">Aprobat</th>' in index
    assert '<th class="num" data-sort-type="number">Linii</th>' in index


def test_city_page_uses_mii_lei_for_every_absolute_budget_value(tmp_path):
    data_root = tmp_path / "data"
    _write_year(data_root, 2026, converted=True, totals=(1_000_000, 1_000_000))
    analysis_path = data_root / "2026" / "01-alba" / "1017-alba-iulia" / "analysis.json"
    payload = json.loads(analysis_path.read_text())
    payload["infografic"] = {
        "unitate": "mii lei",
        "total_cheltuieli": 1_000_000,
        "venituri": {
            "total": 1_000_000,
            "surse": [
                {"cod": "04.02", "nume": "Venituri proprii", "grup": "proprii", "val": 1_000_000},
            ],
            "acoperire_pct": 100.0,
        },
        "capitole": [{
            "cod": "65.02", "nume": "Învățământ", "val": 1_000_000,
            "func": 700_000, "dezv": 300_000,
            "copii": [{"nume": "Învățământ preșcolar", "val": 400_000}],
        }],
        "sectiuni": {"functionare": 700_000, "dezvoltare": 300_000},
        "trim": {
            "functionare": [175_000] * 4,
            "dezvoltare": [75_000] * 4,
            "venituri": [250_000] * 4,
        },
        "ani": {
            "cheltuieli": [1_000_000, 1_100_000, 1_200_000, 1_300_000],
            "venituri": [1_000_000, 1_100_000, 1_200_000, 1_300_000],
        },
    }
    analysis_path.write_text(json.dumps(payload))

    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")
    page = (out / "city" / "1017.html").read_text()

    assert "Funcționare / dezvoltare (mii lei)" in page
    assert '<div class="v">700.000' in page and "/ 300.000" in page
    assert '<div class="bars" id="ig-ven-bars"></div>' in page
    assert 'id="ig-ven-detail" aria-live="polite"' in page
    assert 'id="ig-detail" aria-live="polite"' in page
    assert 'id="ig-ven"' not in page
    assert 'id="ig-ven-list"' not in page
    assert "ig-grid2" not in page
    assert "% din veniturile planificate" in page
    assert "% din secțiunea de funcționare" in page
    assert "% din secțiunea de dezvoltare" in page
    assert 'aria-pressed="true"' in page
    assert "mil. lei" not in page
    assert "milioane lei" not in page
    assert "/ 1000" not in page
    for chart in (
        "venituri", "cheltuieli", "100-lei", "ritm-trimestrial",
        "estimari-multianuale",
    ):
        assert f'data-chart-quality="{chart}"' in page
    assert page.count("<strong>Acoperire:</strong>") == 5
    assert page.count("<strong>Încredere:</strong>") == 5
    assert "recall-ul PDF nu este măsurat" in page


def test_reference_populates_city_meta(data_root, tmp_path):
    ref = tmp_path / "reference"
    ref.mkdir()
    (ref / "municipii.json").write_text(json.dumps({
        "sursa": {"populatie": {"referinta": "2021-12-01", "url": "https://ins.test"}},
        "municipii": {"1017": {
            "populatie": 64227, "populatie_data": "2021-12-01", "suprafata_km2": 103.65,
        }},
    }))
    (ref / "regions_nuts2024.json").write_text(json.dumps({
        "schema_version": 1,
        "dataset_version": "NUTS 2024",
        "source": {"url": "https://eurostat.test/nuts", "join_keys": ["county_code"]},
        "nuts1": {"RO1": "Macroregiunea Unu"},
        "nuts2": {"RO12": "Centru"},
        "counties": {"01": {
            "name": "Alba", "nuts1_code": "RO1", "nuts2_code": "RO12",
            "nuts3_code": "RO121",
        }},
    }))
    (ref / "inflation_hicp.json").write_text(json.dumps({
        "schema_version": 1,
        "dataset_version": "test",
        "source": {"url": "https://eurostat.test/hicp", "join_keys": ["year"]},
        "observations": {"2026": {"annual_average_rate_pct": 3.0}},
    }))
    corpus = agg.build_aggregate(data_root)
    assert corpus.cities[0].populatie == 64227
    assert corpus.cities[0].populatie_data == "2021-12-01"
    assert corpus.cities[0].suprafata_km2 == 103.65
    assert corpus.cities[0].regional_classification.nuts2_code == "RO12"
    assert corpus.cities[0].regional_classification.nuts3_code == "RO121"
    assert corpus.inflation["2026"].annual_average_rate_pct == 3.0
    assert corpus.augmentation_sources["populatie"]["url"] == "https://ins.test"
    assert corpus.augmentation_sources["regional_classification"]["join_keys"] == [
        "county_code"
    ]
    assert corpus.augmentation_sources["inflation"]["join_keys"] == ["year"]

    row = build_analytics(corpus).year_rows(2026)[0]
    assert row.nuts1_code == "RO1" and row.nuts2_name == "Centru"
    assert row.nuts3_code == "RO121"
    assert row.hicp_annual_average_rate_pct == 3.0
    assert row.inflation_status == "final"

    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")
    page = (out / "city" / "1017.html").read_text()
    assert "SIRUTA 1017" in page
    assert "64.227 locuitori (01.12.2021)" in page and "103.65 km²" in page
    assert "regiunea Centru (NUTS 2 RO12)" in page
    # ordered layout: identity, adoption, conversion quality and downloads,
    # then the tabbed budget / execution views closing the page
    assert page.index("SIRUTA 1017") < page.index("Adoptarea bugetului") \
        < page.index("Calitatea conversiei") < page.index("Descarcă Excel") \
        < page.index("Bugetul și execuția lui") \
        < page.index("Cheltuieli planificate pe capitole")


def test_city_page_year_over_year(tmp_path):
    data_root = tmp_path / "data"
    _write_year(data_root, 2025, converted=True, totals=(100.0, 80.0))
    _write_year(data_root, 2026, converted=True, totals=(110.0, 76.0))
    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")

    page = (out / "city" / "1017.html").read_text()
    assert "Evoluție an-la-an" in page
    assert "+10.0%" in page and "-5.0%" in page
    assert '<a href="/repo/2025/city/1017.html">2025</a>' in page
    assert "<strong>2026</strong>" in page

    page25 = (out / "2025" / "city" / "1017.html").read_text()
    assert "Evoluție an-la-an" in page25
    assert '<a href="/repo/city/1017.html">2026</a>' in page25
    assert "<strong>2025</strong>" in page25


def test_analytics_keeps_exclusions_visible_and_writes_auditable_outputs(data_root, tmp_path):
    ref = tmp_path / "reference"
    ref.mkdir()
    (ref / "municipii.json").write_text(json.dumps({
        "sursa": {"populatie": {
            "referinta": "2021-12-01", "url": "https://ins.test/rpl2021",
        }},
        "municipii": {"1017": {
            "populatie": 100, "populatie_data": "2021-12-01", "suprafata_km2": 4,
        }},
    }))
    dataset = build_analytics(agg.build_aggregate(data_root))
    current = next(row for row in dataset.rows if row.year == 2026)
    pending = next(row for row in dataset.rows if row.year == 2025)

    assert current.plan_comparison_eligible is True
    assert current.planned_expense_lei_per_capita == 1200
    assert current.plan_expense_per_capita_rank == 1
    assert current.plan_expense_rank_cohort == 1
    assert current.density_per_km2 == 25
    assert pending.plan_comparison_eligible is False
    assert pending.plan_exclusion_reason == "status_pending"
    assert len(dataset.chapters) == 1

    outputs = write_outputs(dataset, tmp_path / "analytics")
    assert set(outputs) == {"json", "csv", "xlsx"}
    payload = json.loads(outputs["json"].read_text())
    assert payload["coverage"]["2026"]["plan_comparison_eligible"] == 1

    from openpyxl import load_workbook

    workbook = load_workbook(outputs["xlsx"], data_only=False)
    assert workbook.sheetnames == ["Sumar", "Municipii", "Capitole", "Surse augmentare"]
    assert workbook["Municipii"]["R2"].value == \
        '=IF(OR($E2<>"da",O2=""),"",IFERROR(O2*1000/J2,""))'
    assert workbook["Municipii"]["E2"].value == "da"
    source_values = [cell.value for cell in workbook["Surse augmentare"]["C"]]
    assert "https://ins.test/rpl2021" in source_values

    out = tmp_path / "site"
    build_all(data_root, out, base_url="/repo")
    comparison = (out / "comparatii.html").read_text()
    assert "Cheltuieli planificate pe locuitor" in comparison
    assert "1 din" in comparison
    assert (out / "data" / "analytics.xlsx").exists()


def test_incomplete_plan_is_not_ranked(data_root):
    corpus = agg.build_aggregate(data_root)
    city = corpus.cities[0]
    city.populatie = 100
    cy = city.years["2026"]
    cy.executie = agg.Executie(
        trimestru=2, venituri=1000, cheltuieli=900, plan_incomplet=True,
    )
    row = build_analytics(corpus).year_rows(2026)[0]
    assert row.plan_comparison_eligible is False
    assert row.plan_exclusion_reason == "plan_incomplet_fata_de_executie"
    assert row.planned_balance_mii_lei is None
    assert row.planned_expense_lei_per_capita is None
    assert row.plan_expense_per_capita_rank is None
    # The official execution remains independently comparable.
    assert row.execution_comparison_eligible is True
    assert row.actual_expense_lei_per_capita == 9000


def test_public_population_reference_uses_one_census_cohort():
    from pathlib import Path

    payload = json.loads(
        (Path(__file__).parents[1] / "reference" / "municipii.json").read_text()
    )
    municipalities = payload["municipii"]
    assert len(municipalities) == 41
    assert {item["populatie_data"] for item in municipalities.values()} == {"2021-12-01"}
    assert municipalities["60419"]["populatie"] == 263688  # Constanța
    assert municipalities["130534"]["populatie"] == 180540  # Ploiești
    assert municipalities["143450"]["populatie"] == 134309  # Sibiu
    assert municipalities["146263"]["populatie"] == 84322  # Suceava
    assert "Tabel-1.22.xlsx" in payload["sursa"]["populatie"]["fisier_url"]


def test_public_regional_and_inflation_references_are_complete():
    from pathlib import Path

    root = Path(__file__).parents[1]
    manifest = json.loads((root / "data" / "2026" / "manifest.json").read_text())
    county_names = {
        entry["county_code"]: entry["county_name"]
        for entry in manifest["entries"]
    }
    regions = json.loads((root / "reference" / "regions_nuts2024.json").read_text())
    assert regions["dataset_version"] == "NUTS 2024"
    assert set(regions["counties"]) == set(county_names)
    assert {
        code: item["name"] for code, item in regions["counties"].items()
    } == county_names
    assert len({item["nuts3_code"] for item in regions["counties"].values()}) == 42
    assert regions["source"]["join_keys"] == ["county_code"]
    assert regions["source"]["license"]

    inflation = json.loads((root / "reference" / "inflation_hicp.json").read_text())
    assert inflation["source"]["join_keys"] == ["year"]
    assert inflation["source"]["license"]
    assert inflation["observations"]["2024"]["annual_average_rate_pct"] == 5.8
    assert inflation["observations"]["2025"]["annual_average_rate_pct"] == 6.8
    assert "2026" not in inflation["observations"]  # no full-year observation yet


def test_bucharest_identity_wins_over_the_ilfov_manifest_alias(data_root):
    manifest_path = data_root / "2026" / "manifest.json"
    payload = json.loads(manifest_path.read_text())
    source = payload["entries"][0]
    payload["entries"] = [
        {
            **source,
            "county_code": "25", "county_name": "Ilfov",
            "capital_siruta": "179132", "capital_name": "București",
        },
        {
            **source,
            "county_code": "42", "county_name": "București",
            "capital_siruta": "179132", "capital_name": "București",
        },
    ]
    manifest_path.write_text(json.dumps(payload))

    (data_root.parent / "reference").mkdir()
    (data_root.parent / "reference" / "regions_nuts2024.json").write_text(json.dumps({
        "schema_version": 1,
        "dataset_version": "NUTS 2024",
        "source": {},
        "nuts1": {"RO3": "Macroregiunea Trei"},
        "nuts2": {"RO32": "București-Ilfov"},
        "counties": {
            "25": {"name": "Ilfov", "nuts1_code": "RO3", "nuts2_code": "RO32", "nuts3_code": "RO322"},
            "42": {"name": "București", "nuts1_code": "RO3", "nuts2_code": "RO32", "nuts3_code": "RO321"},
        },
    }))

    city = next(
        item for item in agg.build_aggregate(data_root).cities
        if item.siruta == "179132"
    )
    assert city.county == "București" and city.county_code == "42"
    assert city.regional_classification.nuts3_code == "RO321"


def test_execution_rank_uses_one_reporting_period(data_root):
    corpus = agg.build_aggregate(data_root)
    first = corpus.cities[0]
    first.populatie = 100
    first.years["2026"].executie = agg.Executie(
        trimestru=2, la_data="2026-06-30", venituri=1000, cheltuieli=900,
    )
    older = first.model_copy(deep=True)
    older.siruta = "2"
    older.name = "Alt municipiu"
    older.years["2026"].executie = agg.Executie(
        trimestru=1, la_data="2026-03-31", venituri=500, cheltuieli=400,
    )
    corpus.cities.append(older)

    rows = {row.siruta: row for row in build_analytics(corpus).year_rows(2026)}
    assert rows[first.siruta].execution_comparison_eligible is True
    assert rows[first.siruta].execution_expense_per_capita_rank == 1
    assert rows["2"].execution_comparison_eligible is False
    assert rows["2"].execution_exclusion_reason == "perioada_executie_necomparabila"
    assert rows["2"].execution_expense_per_capita_rank is None


def test_real_year_over_year_change_requires_observed_inflation(tmp_path):
    data_root = tmp_path / "data"
    _write_year(data_root, 2025, converted=True, totals=(100.0, 100.0))
    _write_year(data_root, 2026, converted=True, totals=(121.0, 110.0))
    reference = tmp_path / "reference"
    reference.mkdir()
    (reference / "municipii.json").write_text(json.dumps({
        "sursa": {},
        "municipii": {"1017": {"populatie": 100, "populatie_data": "2021-12-01"}},
    }))
    (reference / "inflation_hicp.json").write_text(json.dumps({
        "schema_version": 1,
        "dataset_version": "test",
        "source": {"join_keys": ["year"]},
        "observations": {"2026": {"annual_average_rate_pct": 10.0}},
    }))

    current = build_analytics(agg.build_aggregate(data_root)).year_rows(2026)[0]
    assert current.planned_revenue_yoy_pct == 21.0
    assert current.planned_revenue_yoy_real_pct == 10.0
    assert current.planned_expense_yoy_pct == 10.0
    assert current.planned_expense_yoy_real_pct == 0.0
