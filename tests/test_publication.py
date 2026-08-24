import json
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from bgconvertor.aggregate import build_aggregate
from bgconvertor.export import export as export_workbook
from bgconvertor.manifest import Manifest
from bgconvertor.model import BudgetDocument, BudgetLine, ConversionResult, Issue
from bgconvertor.publication import (
    audit_city,
    audit_data,
    audit_report,
    migration_candidates,
    publish_corpus_result,
)
from bgconvertor.runstore import file_sha256


def _tree(tmp_path):
    data = tmp_path / "data"
    city = data / "2026" / "01-alba" / "1017-alba-iulia"
    city.mkdir(parents=True)
    pdf = city / "budget_file.pdf"
    pdf.write_bytes(b"%PDF-publication-fixture")
    manifest_path = data / "2026" / "manifest.json"
    manifest_path.write_text(json.dumps({
        "year": 2026,
        "entries": [{
            "county_code": "01", "county_name": "Alba",
            "capital_siruta": "1017", "capital_name": "Alba Iulia",
            "path": "01-alba/1017-alba-iulia/budget_file.pdf",
        }],
    }))
    return data, pdf, Manifest(manifest_path)


def _result(*, complete=True):
    clean = BudgetLine(
        code="04.02", name="Cote", kind="revenue", page=1,
        values={"total": Decimal("100"), "trim1": Decimal("25")},
    )
    flagged = BudgetLine(
        code="42.02", name="Subventii", kind="revenue", page=1,
        values={"total": Decimal("50")},
        issues=[Issue(
            check="V2_name", severity="info", message="needs review", page=1,
            code="42.02",
        )],
    )
    doc = BudgetDocument(
        title="BUGET LOCAL", budget="local", suffix="02", pages=[1],
        lines=[clean, flagged],
    )
    return ConversionResult(
        pdf="budget_file.pdf", documents=[doc], pages_expected=1,
        pages_selected=[1], pages_processed=[1] if complete else [],
    )


def test_quality_metric_is_strict_and_does_not_claim_recall():
    stats = _result().stats()
    assert stats["metric"] == "observed_strict_line_rate"
    assert stats["recall_measured"] is False
    assert stats["lines_strictly_verified"] == 1
    assert stats["pct_lines_strictly_verified"] == 50.0
    assert stats["numeric_cells"] == 3
    assert stats["numeric_cells_strictly_verified"] == 2
    assert stats["scope"]["complete_pdf"] is True


def test_numeric_quality_counts_exported_values_on_marker_rows():
    result = _result()
    result.documents[0].lines.insert(0, BudgetLine(
        raw_code="*", code=None, name="TOTAL VENITURI", kind="heading", page=1,
        values={"total": Decimal("150")},
    ))
    stats = result.stats()
    assert stats["quality_schema_version"] == 2
    assert stats["lines"] == 2
    assert stats["numeric_cells"] == 4
    assert stats["numeric_cells_strictly_verified"] == 3


def test_publish_writes_one_hashed_bundle_and_audit_accepts_it(tmp_path):
    data, pdf, manifest = _tree(tmp_path)
    workbook = pdf.with_suffix(".xlsx")
    published = publish_corpus_result(
        _result(), pdf, workbook, manifest,
        llm_preset="test:model", llm_cost_usd=3.25,
    )

    on_disk = Manifest(manifest.path)
    city = on_disk.cities()[0]
    conv = city.entry["conversion"]
    assert conv["status"] == "converted"
    assert conv["quality"]["metric"] == "observed_strict_line_rate"
    assert conv["quality"]["recall_measured"] is False
    assert conv["llm_cost_usd"] == 3.25
    assert conv["artifacts"]["workbook"]["sha256"] == file_sha256(workbook)
    assert conv["artifacts"]["analysis"]["sha256"] == file_sha256(
        pdf.with_name("analysis.json")
    )

    analysis = json.loads(pdf.with_name("analysis.json").read_text())
    bundle = conv["artifacts"]["bundle_id"]
    assert analysis["publication"]["bundle_id"] == bundle
    assert analysis["quality"]["pct_lines_strictly_verified"] == 50.0
    wb = load_workbook(workbook, read_only=True)
    summary = {row[0]: row[1] for row in wb["Sumar calitate"].iter_rows(values_only=True)}
    wb.close()
    assert summary["Bundle conversie"] == bundle
    assert summary["Recall masurat"] == "nu"

    audit = audit_city(on_disk, city)
    assert audit.status == "verified" and audit.trusted
    assert migration_candidates(data) == []
    assert len(migration_candidates(data, include_verified=True)) == 1
    corpus = build_aggregate(data)
    year = corpus.cities[0].years["2026"]
    assert year.status == "converted" and year.artifact_status == "verified"
    assert year.files.xlsx and published["analysis"].exists()


def test_tampering_is_detected_and_not_exposed_to_analytics(tmp_path):
    data, pdf, manifest = _tree(tmp_path)
    publish_corpus_result(_result(), pdf, pdf.with_suffix(".xlsx"), manifest)
    analysis_path = pdf.with_name("analysis.json")
    analysis = json.loads(analysis_path.read_text())
    analysis["quality"]["lines"] = 999
    analysis_path.write_text(json.dumps(analysis))

    fresh = Manifest(manifest.path)
    audit = audit_city(fresh, fresh.cities()[0])
    assert audit.status == "inconsistent"
    assert {issue.code for issue in audit.issues} >= {"metric_mismatch", "hash_mismatch"}
    candidates = migration_candidates(data)
    assert len(candidates) == 1
    assert candidates[0].pdf == pdf
    assert candidates[0].previous_artifact_status == "inconsistent"

    year = build_aggregate(data).cities[0].years["2026"]
    assert year.status == "artifact_mismatch"
    assert not year.has_analysis and year.files.xlsx is None

    report = audit_report(audit_data(data))
    assert report["summary"] == {
        "entries": 1,
        "converted": 1,
        "trusted": 0,
        "inconsistent": 1,
        "by_status": {"inconsistent": 1},
    }


def test_failed_manifest_commit_restores_previous_artifacts(tmp_path, monkeypatch):
    _, pdf, manifest = _tree(tmp_path)
    workbook = pdf.with_suffix(".xlsx")
    analysis = pdf.with_name("analysis.json")
    workbook.write_bytes(b"old-workbook")
    analysis.write_bytes(b"old-analysis")

    def fail(*args, **kwargs):
        raise OSError("simulated manifest failure")

    monkeypatch.setattr(manifest, "set_status", fail)
    with pytest.raises(OSError, match="simulated"):
        publish_corpus_result(_result(), pdf, workbook, manifest)
    assert workbook.read_bytes() == b"old-workbook"
    assert analysis.read_bytes() == b"old-analysis"


def test_incomplete_pdf_cannot_be_published(tmp_path):
    _, pdf, manifest = _tree(tmp_path)
    with pytest.raises(ValueError, match="complete-PDF"):
        publish_corpus_result(_result(complete=False), pdf, pdf.with_suffix(".xlsx"), manifest)
    assert not pdf.with_suffix(".xlsx").exists()
    assert not pdf.with_name("analysis.json").exists()


def test_public_bundle_enforces_five_dollar_external_cost_cap(tmp_path):
    _, pdf, manifest = _tree(tmp_path)
    with pytest.raises(ValueError, match=r"\$5\.00"):
        publish_corpus_result(
            _result(), pdf, pdf.with_suffix(".xlsx"), manifest, llm_cost_usd=5.01
        )


def test_pdf_text_starting_with_equals_is_exported_as_literal_text(tmp_path):
    result = _result()
    result.documents[0].lines.insert(0, BudgetLine(
        code=None,
        name="===== SECTIUNEA TOTAL =====",
        kind="heading",
        page=1,
    ))
    workbook = tmp_path / "literal.xlsx"
    export_workbook(result, workbook)

    wb = load_workbook(workbook, data_only=False, read_only=True)
    assert wb["BL Date"]["C2"].value == "===== SECTIUNEA TOTAL ====="
    assert wb["BL Date"]["C2"].data_type == "s"
    wb.close()
